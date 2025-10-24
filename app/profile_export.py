import os
from typing import Dict, Any, List
from openpyxl import Workbook, load_workbook


SCHEMA_PROMPT_ENGINE: List[str] = [
    "Name", "Gender", "Sexuality", "Age", "Height", "Location",
    "Ethnicity", "Children", "Family plans", "Covid Vaccine", "Pets",
    "Zodiac Sign", "Job title", "University", "Religious Beliefs",
    "Home town", "Politics", "Languages spoken", "Dating Intentions",
    "Relationship type", "Drinking", "Smoking", "Marijuana", "Drugs",
    "prompt_1", "answer_1", "prompt_2", "answer_2", "prompt_3", "answer_3",
    "Other text on profile not covered by above"
]


class ProfileExporter:
    """
    New Excel exporter aligned with the JSON schema from prompt_engine.
    - Creates or appends to profiles.xlsx at the repo root.
    - Flattens 'Profile Prompts and Answers' into six columns.
    - Preserves capitalization and field order from the JSON.
    """

    def __init__(self, export_dir: str, session_id: str, export_xlsx: bool = True) -> None:
        self.export_dir = export_dir
        self.session_id = session_id
        self.export_xlsx = export_xlsx
        self.schema = SCHEMA_PROMPT_ENGINE

        repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
        self.xlsx_path = os.path.join(repo_root, "profiles.xlsx")

        if self.export_xlsx:
            self._ensure_header()

    def _ensure_header(self) -> None:
        """Ensure the Excel file exists and has the correct header."""
        if not os.path.exists(self.xlsx_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "profiles"
            ws.append(self.schema)
            wb.save(self.xlsx_path)
            wb.close()
        else:
            wb = load_workbook(self.xlsx_path)
            ws = wb.active
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if header != self.schema:
                ws.delete_rows(1, ws.max_row)
                ws.append(self.schema)
            wb.save(self.xlsx_path)
            wb.close()

    def append_row(self, profile: Dict[str, Any]) -> None:
        """Append a single profile row to the Excel file."""
        if not self.export_xlsx:
            return

        # Flatten the JSON structure
        row_data = self._flatten_profile(profile)

        # Ensure all schema fields exist
        row = [row_data.get(field, "") for field in self.schema]

        try:
            wb = load_workbook(self.xlsx_path)
            ws = wb.active
            ws.append(row)
            wb.save(self.xlsx_path)
            wb.close()
            print(f"✅ Appended profile to {self.xlsx_path}")
        except Exception as e:
            print(f"❌ Failed to append profile: {e}")

    def _flatten_profile(self, profile: Dict[str, Any]) -> Dict[str, Any]:
        """Flatten nested JSON fields into a flat dict matching the schema."""
        flat: Dict[str, Any] = {}

        for key, value in profile.items():
            if key == "Profile Prompts and Answers" and isinstance(value, list):
                for i, item in enumerate(value[:3], start=1):
                    flat[f"prompt_{i}"] = item.get("prompt", "")
                    flat[f"answer_{i}"] = item.get("answer", "")
            else:
                flat[key] = value

        # Ensure all prompt/answer fields exist even if missing
        for i in range(1, 4):
            flat.setdefault(f"prompt_{i}", "")
            flat.setdefault(f"answer_{i}", "")

        return flat

    def get_paths(self) -> Dict[str, str]:
        return {"xlsx": self.xlsx_path if self.export_xlsx else ""}

    def close(self) -> None:
        pass
